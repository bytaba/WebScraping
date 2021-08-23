from re import X
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def gather ( currency ):
    url = "https://coinmarketcap.com/currencies/%s/" % currency
    data = requests.get(url )
    soup = BeautifulSoup(data.text , 'html.parser')
    price = soup.find('div' , attrs={"class" : "priceValue"})
    return price.text.replace("$" , "")

def entry ( price , row , path):
    book = load_workbook(path)
    sheet = book.active
    sheet.cell(row ,6).value = price
    book.save(path)

def read(path):
    book = load_workbook(path)
    sheet = book.active
    i = 3
    while True :
        name = sheet.cell( i , 2).value
        if name == None:
            break
        print(name)
        entry(gather(name) , i , path)
        i = i + 1

path = "../../Ali.xlsx"
read(path)






